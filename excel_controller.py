from abc import ABC
from abc import abstractmethod

import openpyxl
from pydub.utils import mediainfo

from metadata import MetaData
from metadata import XlsMeta


class ExcelUtils(ABC):

    @abstractmethod
    def get_file_metadata(self, dir_path: str) -> MetaData:
        """
        get a mp3 file metadata
        :param: dir_path as string type
        :return metadata as type MetaData

        """
        NotImplemented

    @abstractmethod
    def read_xlsx_file(self) -> list:
        """
        read a xlsx file
        :param: file_dir as string type (file path)
        :return list of rows element in a worksheet
        """
        NotImplemented


class ExcelFileHandler(ExcelUtils):

    """
    this class with describe how to fill an excel file
    for beelive
    """

    def __init__(self, file_dir: str = None) -> None:
        self.file_dir = file_dir
        self.values: list = []
        self.contrib_data: list = []
        (
            self.workbook,
            self.worksheet,
            self.contributor_worksheet
        ) = self.load_xls_file()

    def add_value(self, *args) -> None:
        for value in args:
            self.values.append(value)

    def add_contrib_values(self, *args) -> None:
        for value in args:
            self.contrib_data.append(value)

    def save_file(self) -> None:
        self.workbook.save("static/METADATA-NEW-TEMPLATE-FR.xltx")

    def load_xls_file(self) -> tuple:
        workbook = openpyxl.load_workbook(self.file_dir)
        workbook.template = True
        worksheet = workbook["Metadatas"]
        contributor_worksheet = workbook.worksheets[1]
        return workbook, worksheet , contributor_worksheet

    def get_metadata(self, metadata: XlsMeta):
        data, contrib_data = [], []
        data.append(metadata.track_number)
        data.append(metadata.track_title)
        data.append(metadata.subtitle)
        data.append(metadata.cd_number)
        data.append(metadata.release_title)
        data.append(metadata.label)
        data.append(metadata.production_year)
        data.append(metadata.production_owner)
        data.append(metadata.copyright_owner)
        data.append(metadata.genre)
        data.append(metadata.sub_genre)
        data.append(metadata.tracktype)
        data.append(metadata.lyrics_language)
        data.append(metadata.title_language)
        data.append(metadata.parental_advisory)
        data.append(metadata.territorie_deliver)
        data.append(metadata.release_price_tier)
        data.append(metadata.track_price_tier)
        data.append(metadata.digital_release_date)
        data.append(metadata.physical_release_date)
        data.append(metadata.simple_start_index)
        data.append(metadata.isrc)
        data.append(metadata.upc_code)
        data.append(metadata.ean_code)
        data.append(metadata.grid)
        data.append(metadata.release_catalog_number)
        data.append(metadata.track_catalog_number)
        data.append(metadata.commercial_desc_en)
        data.append(metadata.commercial_desc_fr)
        data.append(metadata.commercial_desc_gr)
        data.append(metadata.commercial_desc_it)
        data.append(metadata.commercial_desc_sp)

        contrib_data.append(metadata.contributor.contributor_name)
        contrib_data.append(metadata.contributor.role1)
        contrib_data.append(metadata.contributor.role2)
        contrib_data.append(metadata.contributor.role3)
        contrib_data.append(metadata.contributor.role4)
        contrib_data.append(metadata.contributor.release_title)
        contrib_data.append(metadata.contributor.track)
        contrib_data.append(metadata.contributor.spotify_id)
        contrib_data.append(metadata.contributor.apple_music_id)
        return data, contrib_data

    def get_file_metadata(self, dir_path: str) -> MetaData:
        media_data: dict = dict()
        if dir_path.endswith(".mp3"):
            media_data = mediainfo(dir_path)
        return MetaData(song_data=media_data)


    def write_metadata_to_excel(self, sheet, data, start_row, workbook):
        for row_index, row_data in enumerate(data, start=start_row):
            for col_index, value in enumerate(row_data, start=1):
                c1 = sheet.cell(row=row_index, column=col_index)
                c1.value = value
        workbook.save("static/METADATA-NEW-TEMPLATE-FR.xltx")

    def write_contributeur_to_excel(self, sheet, data, start_row, workbook):
        for row_index, row_data in enumerate(data, start=start_row):
            for col_index, value in enumerate(row_data, start=1):
                c1 = sheet.cell(row=row_index, column=col_index)
                c1.value = value
        workbook.save("static/METADATA-NEW-TEMPLATE-FR.xltx")

    def write_in_xlsx_file(self, data, meta_start=3, contrib_start=2) -> None:
        self.write_metadata_to_excel(self.worksheet, data[0], meta_start,  self.workbook)
        self.write_contributeur_to_excel(self.contributor_worksheet, data[1], contrib_start, self.workbook)
        self.values = []
        self.contrib_data = []


    def read_xlsx_file(self) -> list:
        return [list(row) for row in self.worksheet.values]

    def read_xlsx_files(self) -> list:
        return [list(row) for row in self.contributor_worksheet.values]


if __name__ == "__main__":
    pass
