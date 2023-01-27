import openpyxl


class File:
    def __init__(self, file_path: str) -> None:
        self.file_path = file_path

    def modify_excel_file(self):
        wb = openpyxl.load_workbook(self.file_path)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=85, max_row=88, min_col=1, max_col=5):
            for cell in row:
                print(cell.value)
        wb.close()

    def write_excel_file(self):
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet()
        for i in range(100):
            ws.append(['%d' % i for i in range(200)])
        wb.save("D:\output\michier1.xlsx")


if __name__ == "__main__":
    # File("D:\output\michier.xlsx").modify_excel_file()
    # File(file_path=None).write_excel_file()
    from multiprocessing import cpu_count
    from multiprocessing.dummy import Pool
    # example of parallel map_async() with the thread pool
    from random import random
    from time import sleep
    from multiprocessing.pool import ThreadPool

    def task(identifier):
        # generate a value
        value = random()
        print(f'Task {identifier} executing with {value}')
        sleep(value)
        return value

    # with ThreadPool() as pool:
    #     result = pool.map_async(task, range(10))
    #     for result in result.get():
    #         print(f'Got result: {result}')

    import asyncio
    import threading

    from datetime import datetime
    now = datetime.now

    async def test_timer_function():
        await asyncio.sleep(2)
        print(f"ending async task at {now()}")
        return

    def run_async_loop_in_thread():
        asyncio.run(test_timer_function())

    def main():
        print(f"Starting timer at {now()}")
        t = threading.Thread(target=run_async_loop_in_thread)
        t.start()
        print(f"Ending timer at {now()}")
        return t

    t = main()
    t.join()
    print(f"asyncio thread exited normally at {now()}")
